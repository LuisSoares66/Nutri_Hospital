import os
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
import load_catalogo_produtos_from_excel

def _path(data_dir: str, filename: str) -> str:
    return os.path.join(data_dir, filename)


def _norm_header(v: Any) -> str:
    """Normaliza nome de coluna."""
    if v is None:
        return ""
    return str(v).strip()


def _cell_to_str(v: Any) -> str:
    """Converte valores de célula para string (sem None)."""
    if v is None:
        return ""
    # números grandes podem vir como float: 4.0 -> "4"
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _cell_to_int(v: Any) -> Optional[int]:
    """Converte célula para int se possível."""
    if v is None or str(v).strip() == "":
        return None
    try:
        # trata "4", "4.0", 4, 4.0
        return int(float(str(v).strip()))
    except Exception:
        return None


def _read_sheet_rows(
    wb_path: str,
    sheet_name: Optional[str] = None,
    header_row: int = 1,
) -> List[Dict[str, Any]]:
    """
    Lê uma planilha e retorna lista de dicts:
    [{coluna: valor, ...}, ...]
    """
    if not os.path.exists(wb_path):
        return []

    wb = load_workbook(wb_path, data_only=True)

    ws = wb[sheet_name] if sheet_name else wb.active

    # headers
    headers = []
    for cell in ws[header_row]:
        headers.append(_norm_header(cell.value))

    # ignora colunas vazias no final
    # (mantém índice para mapear certinho)
    rows: List[Dict[str, Any]] = []
    for r in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if r is None:
            continue
        # pula linha totalmente vazia
        if all((c is None or str(c).strip() == "") for c in r):
            continue

        obj: Dict[str, Any] = {}
        for idx, h in enumerate(headers):
            if not h:
                continue
            val = r[idx] if idx < len(r) else None
            obj[h] = val if val is not None else ""

        rows.append(obj)

    return rows


# ======================================================
# HOSPITAIS
# ======================================================
def load_hospitais_from_excel(data_dir: str) -> List[Dict[str, Any]]:
    """
    hospitais.xlsx colunas:
    - id_hospital
    - nome_hospital
    - endereco
    - numero
    - complemento
    - cep
    - cidade
    - estado
    """
    fp = _path(data_dir, "hospitais.xlsx")
    rows = _read_sheet_rows(fp)

    out: List[Dict[str, Any]] = []
    for r in rows:
        out.append({
            "id_hospital": _cell_to_int(r.get("id_hospital")),
            "nome_hospital": _cell_to_str(r.get("nome_hospital")),
            "endereco": _cell_to_str(r.get("endereco")),
            "numero": _cell_to_str(r.get("numero")),
            "complemento": _cell_to_str(r.get("complemento")),
            "cep": _cell_to_str(r.get("cep")),
            "cidade": _cell_to_str(r.get("cidade")),
            "estado": _cell_to_str(r.get("estado")),
        })
    return out


# ======================================================
# CONTATOS
# ======================================================
def load_contatos_from_excel(data_dir: str) -> List[Dict[str, Any]]:
    """
    contatos.xlsx colunas:
    - id_contato
    - id_hospital
    - hospital_nome
    - nome_contato
    - cargo
    - telefone
    """
    fp = _path(data_dir, "contatos.xlsx")
    rows = _read_sheet_rows(fp)

    out: List[Dict[str, Any]] = []
    for r in rows:
        out.append({
            "id_contato": _cell_to_int(r.get("id_contato")),
            "id_hospital": _cell_to_int(r.get("id_hospital")),
            "hospital_nome": _cell_to_str(r.get("hospital_nome")),
            "nome_contato": _cell_to_str(r.get("nome_contato")),
            "cargo": _cell_to_str(r.get("cargo")),
            "telefone": _cell_to_str(r.get("telefone")),
        })
    return out


# ======================================================
# DADOS DO HOSPITAL
# ======================================================
def load_dados_hospitais_from_excel(data_dir: str) -> List[Dict[str, Any]]:
    """
    dadoshospitais.xlsx colunas:
    - id_hospital
    - Qual a especialidade do hospital?
    - Quantos leitos?
    - Quantos leitos de UTI?
    - ...
    """
    fp = _path(data_dir, "dadoshospitais.xlsx")
    rows = _read_sheet_rows(fp)

    # Mantém os nomes de colunas exatamente como estão na planilha
    # e só garante o id_hospital como int.
    out: List[Dict[str, Any]] = []
    for r in rows:
        rr = dict(r)
        rr["id_hospital"] = _cell_to_int(r.get("id_hospital"))
        # normaliza Nones para ""
        for k, v in list(rr.items()):
            if v is None:
                rr[k] = ""
        out.append(rr)

    return out


# ======================================================
# PRODUTOS POR HOSPITAL
# ======================================================
def load_produtos_hospitais_from_excel(data_dir: str) -> List[Dict[str, Any]]:
    """
    produtoshospitais.xlsx colunas:
    - hospital_id  (ou id_hospital dependendo da sua planilha)
    - nome_hospital
    - produto
    - quantidade
    (pode ter marca_planilha se você adicionou)
    """
    fp = _path(data_dir, "produtoshospitais.xlsx")
    rows = _read_sheet_rows(fp)

    out: List[Dict[str, Any]] = []
    for r in rows:
        # aceita hospital_id ou id_hospital
        hid = _cell_to_int(r.get("hospital_id"))
        if hid is None:
            hid = _cell_to_int(r.get("id_hospital"))

        qtd = _cell_to_int(r.get("quantidade"))
        if qtd is None:
            # se vier "2.0"
            try:
                qtd = int(float(str(r.get("quantidade") or "0").strip()))
            except Exception:
                qtd = 0

        out.append({
            "hospital_id": hid,
            "nome_hospital": _cell_to_str(r.get("nome_hospital")),
            "marca_planilha": _cell_to_str(r.get("marca_planilha")),
            "produto": _cell_to_str(r.get("produto")),
            "quantidade": qtd,
        })
    return out


# ======================================================
# CATÁLOGO DE PRODUTOS (produtos.xlsx com várias abas)
# ======================================================
def load_catalogo_produtos(data_dir: str) -> Dict[str, List[Dict[str, Any]]]:
    """
    produtos.xlsx tem abas:
    PRODIET, NESTLÉ, DANONE, FRESENIUS
    cada aba com colunas:
    Produto, Embalagem, Referência, KCAL, PTN (g), LIP (g), FIBRAS (g), ...
    Retorna:
    {"PRODIET":[{...},{...}], "NESTLÉ":[...], ...}
    """
    fp = _path(data_dir, "produtos.xlsx")
    if not os.path.exists(fp):
        return {}

    wb = load_workbook(fp, data_only=True)
    out: Dict[str, List[Dict[str, Any]]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # lê header primeira linha
        headers = [_norm_header(c.value) for c in ws[1]]
        items: List[Dict[str, Any]] = []

        for r in ws.iter_rows(min_row=2, values_only=True):
            if r is None or all((c is None or str(c).strip() == "") for c in r):
                continue

            obj: Dict[str, Any] = {}
            for idx, h in enumerate(headers):
                if not h:
                    continue
                val = r[idx] if idx < len(r) else None
                obj[h] = _cell_to_str(val)

            # garante que tenha "Produto"
            if obj.get("Produto"):
                items.append(obj)

        out[sheet_name] = items

    return out
